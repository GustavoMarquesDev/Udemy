
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORK_BOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# inicia seu workbook
workbook = Workbook()

# pega a planilha ativa sheet por padrao
# worksheet: Worksheet = workbook.active  # type: ignore

# para criar uma planilha diferente e coloca- la no índice 0
worksheet = workbook.create_sheet(title='Students', index=0)

# remover uma planilha
workbook.remove(workbook['Sheet'])

# cria os cabeçalhos das colunas na planilha
worksheet.cell(row=1, column=1, value='Name')
worksheet.cell(row=1, column=2, value='Age')
worksheet.cell(row=1, column=3, value='Grade')

students = [
    ['Joao', 14, 5.5],
    ['Maria', 12, 6.0],
    ['Pedro', 13, 9.0],
    ['Jose', 15, 6.5],
]

# uma maneira mais completa porem mais complicada
# for i, student_row in enumerate(students, start=2):
#     for j, student_column in enumerate(student_row, start=1):
#         worksheet.cell(row=i, column=j, value=student_column)

for student in students:
    worksheet.append(student)

# para salvar o workbook no diretório especificado
workbook.save(WORK_BOOK_PATH)
