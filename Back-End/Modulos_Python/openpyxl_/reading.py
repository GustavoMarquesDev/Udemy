
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORK_BOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# carregando um arquivo do excel
workbook: Workbook = load_workbook(WORK_BOOK_PATH)

sheet_name = 'Students'

worksheet: Worksheet = workbook[sheet_name]

for row in worksheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value, end='\t')

    # para alterar uma celula específica
        if cell.value == 'Maria':
            worksheet.cell(row=cell.row, column=2, value=23)

    print()

# edita uma celula especifica
worksheet['B3'].value = '14'

# para salvar o workbook no diretório especificado
workbook.save(WORK_BOOK_PATH)
